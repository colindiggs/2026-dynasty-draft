"""
Live draft assistant for The Phoenix League.

When the draft is on the clock, run this to see:
  - Picks made so far
  - Your roster after picks
  - Top N players still available, ranked by your My Board (2026_board_data_driven.xlsx)
  - Your next pick number and where it falls in My Board ranks

Usage:
    python live_draft.py            # uses the league draft id
    python live_draft.py <draft_id> # override

Requires: pip install requests openpyxl
"""
from __future__ import annotations
import os
import sys
import time

try:
    import requests
except ImportError:
    print("ERROR: pip install requests")
    sys.exit(1)

from openpyxl import load_workbook

# Defaults
LEAGUE_ID = "1335118685588701184"
DEFAULT_DRAFT_ID = "1335118685605498880"
MY_USER_ID = "1250202714625822720"
MY_ROSTER_ID = 3
BOARD_XLSX = "2026_board_data_driven.xlsx"
SHOW_TOP_N = 30


def get(url, retries=3):
    for i in range(retries):
        try:
            r = requests.get(url, timeout=15)
            if r.status_code == 200:
                return r.json()
            print(f"  [warn] {url} -> HTTP {r.status_code}, retry {i+1}/{retries}")
        except Exception as e:
            print(f"  [warn] {url} -> {e}, retry {i+1}/{retries}")
        time.sleep(1 + i)
    print(f"  [fail] {url} — giving up")
    return None


def load_board():
    """Read My Board tab, return list of {rank, player, pos, team, age, n_src}."""
    if not os.path.exists(BOARD_XLSX):
        print(f"ERROR: {BOARD_XLSX} not found. Run build_xlsx_v2.py first.")
        sys.exit(1)
    wb = load_workbook(BOARD_XLSX, data_only=True)
    ws = wb["My Board"]
    rows = list(ws.iter_rows(min_row=4, values_only=True))
    headers = rows[0]
    # Find column indices
    idx = {h: i for i, h in enumerate(headers) if h}
    out = []
    for r in rows[1:]:
        if not r[idx["My Rk"]]:
            continue
        out.append({
            "rank": r[idx["My Rk"]],
            "tier": r[idx["Tier"]],
            "player": r[idx["Player"]],
            "pos": r[idx["Pos"]],
            "team": r[idx["Team"]],
            "age": r[idx["Age"]],
            "n_src": r[idx.get("n Src", -1)] if "n Src" in idx else None,
            "fit": r[idx.get("Fit", -1)] if "Fit" in idx else "",
            "notes": r[idx.get("Notes", -1)] if "Notes" in idx else "",
        })
    return out


def norm(s):
    if not s:
        return ""
    s = s.strip().lower().replace(".", "").replace(",", "")
    import re
    s = re.sub(r"\s+(jr|sr|ii|iii|iv)\b", "", s)
    s = re.sub(r"[^a-z0-9 ']+", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def main():
    draft_id = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_DRAFT_ID

    print(f"Fetching draft {draft_id}...")
    meta = get(f"https://api.sleeper.app/v1/draft/{draft_id}")
    if meta is None:
        print("Draft fetch failed (Sleeper API blip or wrong ID). Stopping.")
        sys.exit(1)
    picks = get(f"https://api.sleeper.app/v1/draft/{draft_id}/picks") or []
    print(f"  status={meta.get('status')}  picks_made={len(picks)}  type={meta.get('type')}")

    pidx_path = os.path.join("sources", "_players_idx.json")
    if not os.path.exists(pidx_path):
        print("Fetching Sleeper player index (one-time, ~5MB)...")
        pidx = get("https://api.sleeper.app/v1/players/nfl")
        if pidx is None:
            print("ERROR: player index fetch failed.")
            sys.exit(1)
        os.makedirs("sources", exist_ok=True)
        import json
        with open(pidx_path, "w", encoding="utf-8") as f:
            json.dump(pidx, f)
    else:
        import json
        with open(pidx_path, encoding="utf-8") as f:
            pidx = json.load(f)

    # Picks made -> set of normalized player names
    taken = set()
    pick_log = []
    for pk in picks:
        pid = pk.get("player_id")
        p = pidx.get(pid, {}) if pid else {}
        name = (p.get("full_name") or
                f"{p.get('first_name','')} {p.get('last_name','')}").strip()
        roster_id = pk.get("roster_id") or pk.get("picked_by")
        is_mine = (roster_id == MY_ROSTER_ID)
        pick_log.append({
            "no": pk.get("pick_no"),
            "round": pk.get("round"),
            "name": name,
            "pos": p.get("position"),
            "team": p.get("team"),
            "is_mine": is_mine,
        })
        if name:
            taken.add(norm(name))

    board = load_board()
    print(f"\nLoaded {len(board)} players from {BOARD_XLSX} My Board.")

    # Show my picks so far
    mine = [pk for pk in pick_log if pk["is_mine"]]
    print(f"\n=== My picks so far: {len(mine)} ===")
    for pk in mine:
        print(f"  {pk['no']:>3} (R{pk['round']}): {pk['name']} {pk['pos']} {pk['team']}")

    # Last 5 picks made (across the board, for context)
    print(f"\n=== Last 5 picks (any team) ===")
    for pk in pick_log[-5:]:
        marker = "***" if pk["is_mine"] else "   "
        print(f"  {marker} {pk['no']:>3} (R{pk['round']}): {pk['name']:25s} {pk['pos']} {pk['team']}")

    # Top N still available, by My Board rank
    available = [b for b in board if norm(b["player"]) not in taken]
    print(f"\n=== Top {SHOW_TOP_N} on My Board still available ===")
    print(f"{'Rk':>4} {'Tier':>4} {'Player':25s} {'Pos':3s} {'Team':4s} {'Age':>3} {'Fit':>4} {'nSrc':>4}")
    for b in available[:SHOW_TOP_N]:
        print(f"{b['rank']:>4} {b['tier']:>4} {b['player']:25s} {b['pos']:3s} "
              f"{(b['team'] or 'FA'):4s} {str(b['age'] or '-'):>3} "
              f"{(b['fit'] or ''):>4} {str(b['n_src'] or '-'):>4}")

    # My next pick: derive from total picks taken + 12-team linear logic
    # Linear: round R, pick P -> overall = (R-1)*12 + P. My slot = 11.
    n_taken = len(picks)
    # Find next pick for slot 11 in each round
    print(f"\n=== Your next picks ===")
    for rd in range(1, 5):
        overall = (rd - 1) * 12 + 11
        if overall > n_taken:
            on_clock_in = overall - n_taken
            print(f"  R{rd}.11 (overall {overall}) — {on_clock_in} picks away")
            break


if __name__ == "__main__":
    main()
