"""
Data-driven dynasty board builder for The Phoenix League (12-team SF / TEP / IDP).

Reads sources/*.csv per sources/sources.json, blends them with a Sleeper-backed
player index for pos/age/team backfill, and writes the xlsx.

Adds vs v1: IDP positions (LB/DL/DB/EDGE/S), Sleeper backfill, Roster Fit column,
SF/TEP/IDP-aware notes.

Run: python build_xlsx_v2.py
"""
from __future__ import annotations
import csv
import json
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============= CONFIG =============

SOURCES_DIR = "sources"
SOURCES_JSON = os.path.join(SOURCES_DIR, "sources.json")
PLAYER_IDX_CACHE = os.path.join(SOURCES_DIR, "_players_idx.json")
NFL_DRAFT_JSON = os.path.join(SOURCES_DIR, "nfl_draft_2026.json")
OUT_XLSX = "2026_board_data_driven.xlsx"
DOCS_DIR = "docs"
OUT_JSON = os.path.join(DOCS_DIR, "data.json")

# League: The Phoenix League (Sleeper league_id 1335118685588701184)
LEAGUE_NAME = "The Phoenix League"
LEAGUE_FORMAT = "12-team / SF / TEP (rec_te=1.0) / IDP / linear"

# Capital Score (algorithmic baseline) — fed from sources/nfl_draft_2026.json.
# Edit that JSON when draft data changes; this dict is just the parsed view.
def _load_nfl_draft():
    if not os.path.exists(NFL_DRAFT_JSON):
        print(f"  [warn] {NFL_DRAFT_JSON} not found — every player will be UDFA")
        return {}
    with open(NFL_DRAFT_JSON, encoding="utf-8") as f:
        raw = json.load(f)
    return {
        name: tuple(v)
        for name, v in raw.items()
        if not name.startswith("_")
    }

NFL_DRAFT = _load_nfl_draft()

# Capital score position bumps. Negative = better (lower cap_score = higher rank).
# IDP is half-points in Phoenix League — IDP rookies get a POSITIVE bump
# (deweighted) so their cap rank doesn't compete with offensive rookies.
# Reference: 1Q-style equivalence — half-pt IDP rookie ~ 50-60% of equivalent
# offensive rookie value, so we add a constant penalty.
POS_BUMP = {
    "QB": -18, "TE": -10, "RB": -6, "WR": 0,
    "DL": +25, "EDGE": +25, "LB": +20, "DB": +30, "S": +30, "CB": +30,
}
ROUND_MULT = {1: 1.0, 2: 1.05, 3: 1.15, 4: 1.25, 5: 1.40, 6: 1.60, 7: 1.80}
UDFA_BASE, UDFA_MULT = 270, 2.00
CAP_WEIGHT = 0.5

# Colin's picks: 1.11, 2.11, 3.09 (from CThompson1512), 3.11, 4.11
PICK_AT_RANK = {
    11: "1.11 Own", 23: "2.11 Own",
    33: "3.09 Reap", 35: "3.11 Own", 47: "4.11 Own",
}

# Roster-fit weights, derived from Colin's actual roster on 2026-04-30.
# IDP positions are CAPPED LOW because Phoenix League scores IDP at half points,
# so an IDP rookie is worth ~half an offensive rookie at the same draft slot.
# "Need" reflects roster gap × scoring importance.
ROSTER_NEED = {
    "WR": 0.7,    # 10 WRs but 5 are 25+. Need young top-end.
    "RB": 0.6,    # Have Henderson; aging Mixon/JT. Always need RBs in dynasty.
    "TE": 0.5,    # Warren+Fannin young, Pitts 25, Okonkwo 26. Marginal need.
    "QB": 0.0,    # Lamar/Purdy/McKee/Dart — no need.
    "DL": 0.4,    # Roster gap, but half-pt IDP caps real value.
    "EDGE": 0.4,
    "LB": 0.3,    # Cooper young; Oluokun aging — depth target only.
    "DB": 0.2,
    "S": 0.2,
    "CB": 0.2,
}


# ============= NAME NORMALIZATION =============

def norm_name(name: str) -> str:
    s = name.strip().lower()
    s = s.replace(".", "").replace(",", "")
    s = re.sub(r"\s+(jr|sr|ii|iii|iv)\b", "", s)
    s = re.sub(r"[^a-z0-9 ']+", "", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


# ============= TEAM ABBREVIATION NORMALIZATION =============

TEAM_ALIASES = {
    "ARZ": "ARI", "LVR": "LV", "OAK": "LV", "CLV": "CLE", "WSH": "WAS",
    "KCC": "KC", "TBB": "TB", "SFO": "SF", "NOS": "NO", "NEP": "NE",
    "JAC": "JAX", "GBP": "GB", "LAR": "LAR", "LAC": "LAC", "BLT": "BAL",
    "HST": "HOU",
}

def norm_team(t: str) -> str:
    if not t:
        return ""
    t = t.strip().upper()
    return TEAM_ALIASES.get(t, t)


# ============= SLEEPER PLAYER INDEX BACKFILL =============

def load_sleeper_idx() -> dict:
    """Load cached Sleeper player index. Returns dict keyed by norm_name."""
    if not os.path.exists(PLAYER_IDX_CACHE):
        print(f"  [warn] {PLAYER_IDX_CACHE} not found. Run sleeper_scraper.py once "
              f"or fc_fetcher.py to populate. Skipping backfill.")
        return {}
    with open(PLAYER_IDX_CACHE, encoding="utf-8") as f:
        raw = json.load(f)
    by_name = {}
    for pid, p in raw.items():
        full = p.get("full_name") or f"{p.get('first_name','')} {p.get('last_name','')}".strip()
        if not full:
            continue
        n = norm_name(full)
        if not n:
            continue
        by_name[n] = {
            "sleeper_id": pid,
            "pos": p.get("position") or "",
            "team": (p.get("team") or "FA"),
            "age": p.get("age"),
            "years_exp": p.get("years_exp"),
            "college": p.get("college") or "",
        }
    return by_name


# ============= LOAD SOURCES =============

def load_sources():
    with open(SOURCES_JSON, encoding="utf-8") as f:
        registry = json.load(f)
    sources = {}
    for key, cfg in registry.items():
        if key.startswith("_"):
            continue
        path = os.path.join(SOURCES_DIR, cfg["csv"])
        if not os.path.exists(path):
            print(f"  [skip] {key} ({cfg['label']}) - {path} missing")
            continue
        ranks = {}
        with open(path, encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                player = (row.get("player") or "").strip()
                if not player:
                    continue
                try:
                    rk = int(row["source_rank"])
                except (KeyError, ValueError, TypeError):
                    continue
                val = None
                for k in ("value", "ktc_value", "fc_value"):
                    if k in row and row[k]:
                        try:
                            val = int(float(row[k]))
                            break
                        except ValueError:
                            pass
                ranks[norm_name(player)] = {
                    "rank": rk, "value": val, "name": player,
                    "pos": row.get("pos") or "",
                    "team": norm_team(row.get("team") or ""),
                }
        sources[key] = {
            "label": cfg["label"], "weight": float(cfg.get("weight", 1.0)),
            "ranks": ranks,
        }
        print(f"  [loaded] {key:12s} ({cfg['label'][:45]:45s}) {len(ranks):>3} players, w={cfg['weight']}")
    return sources


# ============= COMPUTE CAPITAL =============

def cap_score(pos, rd, pk):
    if rd is None:
        return UDFA_BASE * UDFA_MULT + POS_BUMP.get(pos, 0)
    return pk * ROUND_MULT.get(rd, 2.0) + POS_BUMP.get(pos, 0)


# ============= BUILD MASTER PLAYER LIST =============

def build_players(sources, sleeper_idx):
    players = {}

    # Seed from NFL_DRAFT (offensive 2026 picks we have full metadata for)
    for name, (pos, college, team, rd, pk) in NFL_DRAFT.items():
        n = norm_name(name)
        si = sleeper_idx.get(n, {})
        players[n] = {
            "name": name, "pos": pos, "team": team,
            "age": si.get("age"),
            "rd": rd, "pk": pk, "college": college,
            "sleeper_id": si.get("sleeper_id"), "years_exp": si.get("years_exp"),
            "source_ranks": {}, "source_values": {},
        }

    # Add anyone present in any source
    for key, src in sources.items():
        for n, info in src["ranks"].items():
            if n not in players:
                si = sleeper_idx.get(n, {})
                players[n] = {
                    "name": info["name"],
                    "pos": info["pos"] or si.get("pos") or "?",
                    "team": info["team"] or si.get("team") or "FA",
                    "age": si.get("age"),
                    "rd": None, "pk": None, "college": si.get("college"),
                    "sleeper_id": si.get("sleeper_id"),
                    "years_exp": si.get("years_exp"),
                    "source_ranks": {}, "source_values": {},
                }
            else:
                # Backfill team/pos
                if not players[n]["team"] or players[n]["team"] == "FA":
                    if info["team"]:
                        players[n]["team"] = info["team"]
                if (not players[n]["pos"] or players[n]["pos"] == "?") and info["pos"]:
                    players[n]["pos"] = info["pos"]
            players[n]["source_ranks"][key] = info["rank"]
            if info.get("value") is not None:
                players[n]["source_values"][key] = info["value"]

    # Compute Capital Score
    for p in players.values():
        p["cap_score"] = cap_score(p["pos"], p["rd"], p["pk"])

    sorted_cap = sorted(players.values(), key=lambda x: x["cap_score"])
    for i, p in enumerate(sorted_cap, 1):
        p["cap_rk"] = i

    return players


# ============= BLEND =============

def compute_blend(players, sources):
    """
    Weighted-mean blend, but with a missing-source penalty so a player ranked
    high in 1 source can't leapfrog players ranked everywhere. For each source
    a player is NOT in, charge them a penalty rank = total_source_size + 5.
    This keeps the math simple and well-defined and treats "missing" as
    "ranked outside this source's coverage."
    """
    for p in players.values():
        weights_used = []
        ranks_used = []
        n_src = 0
        for key, src in sources.items():
            w = src["weight"]
            if key in p["source_ranks"]:
                ranks_used.append(p["source_ranks"][key])
                weights_used.append(w)
                n_src += 1
            else:
                penalty = len(src["ranks"]) + 5
                ranks_used.append(penalty)
                weights_used.append(w)
        ranks_used.append(p["cap_rk"])
        weights_used.append(CAP_WEIGHT)
        num = sum(r * w for r, w in zip(ranks_used, weights_used))
        den = sum(weights_used)
        p["final"] = num / den
        p["n_sources"] = n_src

    ranked = sorted(players.values(), key=lambda x: x["final"])
    for i, p in enumerate(ranked, 1):
        p["my_rk"] = i

    def tier(rk):
        if rk <= 6: return 1
        if rk <= 12: return 2
        if rk <= 20: return 3
        if rk <= 30: return 4
        if rk <= 45: return 5
        if rk <= 65: return 6
        return 7
    for p in ranked:
        p["tier"] = tier(p["my_rk"])
    return ranked


# ============= ROSTER FIT =============

def roster_fit(p):
    """Score from 0 (skip) to 5 (must-target) based on position need + age."""
    need = ROSTER_NEED.get(p["pos"], 0.3)
    age = p.get("age")
    age_bonus = 0.0
    if age is not None:
        if age <= 21: age_bonus = 0.3
        elif age <= 23: age_bonus = 0.1
        else: age_bonus = -0.2
    score = need + age_bonus
    if score >= 1.0: return ("***", "Must-target")
    if score >= 0.7: return ("**",  "Good fit")
    if score >= 0.4: return ("*",   "Marginal")
    if score >= 0.0: return ("",    "")
    return ("-", "Skip")


# ============= BUILD WORKBOOK =============

HEADER = Font(name="Arial", bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", start_color="1F3A4D")
SUBHEADER = Font(name="Arial", bold=True, size=10)
SUBHEADER_FILL = PatternFill("solid", start_color="D9E2EC")
NORMAL = Font(name="Arial", size=10)
NOTE = Font(name="Arial", italic=True, color="555555", size=9)
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
TIER_FILLS = {
    1: PatternFill("solid", start_color="C8E6C9"),
    2: PatternFill("solid", start_color="DCEDC8"),
    3: PatternFill("solid", start_color="FFF9C4"),
    4: PatternFill("solid", start_color="FFE0B2"),
    5: PatternFill("solid", start_color="FFCCBC"),
    6: PatternFill("solid", start_color="F8BBD0"),
    7: PatternFill("solid", start_color="EEEEEE"),
}


def header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = HEADER
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER


def build_my_board(wb, ranked, sources):
    ws = wb.create_sheet("My Board")
    ws["A1"] = f"My Board — {LEAGUE_NAME} ({LEAGUE_FORMAT})"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    src_keys = list(sources.keys())
    n_src_cols = len(src_keys)
    # static + per-source + ktc_val + cap + spread + n_src + fit + notes
    total_cols = 9 + n_src_cols + 6
    last_letter = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{last_letter}1")
    weights_str = ", ".join(f"{k}={sources[k]['weight']}" for k in src_keys)
    ws["A2"] = (f"My Rk = weighted blend. Weights: {weights_str}, cap={CAP_WEIGHT}. "
                f"Roster Fit: ***=must-target, **=good, *=marginal, -=skip. Spread = max-min source rank.")
    ws["A2"].font = NOTE
    ws.merge_cells(f"A2:{last_letter}2")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.row_dimensions[2].height = 30

    cols = ["My Rk", "Tier", "Slot", "Player", "Pos", "Age", "Team", "Rd", "Pk"]
    cols += [k.upper() for k in src_keys]
    cols += ["KTC Val", "Cap Rk", "Spread", "n Src", "Fit", "Notes"]
    header_row(ws, 4, cols)

    for i, p in enumerate(ranked, 5):
        rd_str = p["rd"] if p["rd"] else ("Drft*" if p["source_ranks"] else "UDFA")
        pk_str = p["pk"] if p["pk"] else "-"
        slot = PICK_AT_RANK.get(p["my_rk"], "")
        ranks_present = list(p["source_ranks"].values())
        spread = max(ranks_present) - min(ranks_present) if len(ranks_present) >= 2 else None

        notes = []
        if p["pos"] == "TE": notes.append("TEP")
        if p["pos"] == "QB": notes.append("SF")
        if p["pos"] in ("DL", "EDGE", "LB", "DB", "S", "CB"): notes.append("IDP")
        if p.get("rd") == 1: notes.append("R1 cap")
        if p.get("rd") and p["rd"] >= 6: notes.append("late dart")
        if not p["source_ranks"]: notes.append("no source")
        # Disagreement signals
        ktc_rk = p["source_ranks"].get("ktc")
        if ktc_rk is not None and abs(ktc_rk - p["cap_rk"]) >= 8:
            notes.append("KTC loves" if ktc_rk < p["cap_rk"] else "KTC fades")
        if spread is not None and spread >= 15:
            notes.append(f"high spread ({spread})")

        fit_mark, _ = roster_fit(p)

        ktc_val = p["source_values"].get("ktc")
        row_data = [
            p["my_rk"], p["tier"], slot, p["name"], p["pos"], p.get("age") or "",
            p["team"] or "FA", rd_str, pk_str,
        ]
        for k in src_keys:
            row_data.append(p["source_ranks"].get(k, ""))
        row_data += [ktc_val or "", p["cap_rk"], spread or "", p.get("n_sources", 0),
                     fit_mark, "; ".join(notes)]

        fill = TIER_FILLS[p["tier"]]
        for j, v in enumerate(row_data, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = NORMAL
            c.fill = fill
            c.border = BORDER
            if j in (3, 4, total_cols):
                c.alignment = Alignment(horizontal="left", wrap_text=(j == total_cols))
            else:
                c.alignment = Alignment(horizontal="center")
        if slot:
            for j in range(1, total_cols + 1):
                ws.cell(row=i, column=j).font = Font(name="Arial", size=10, bold=True)

    widths = [6, 5, 11, 25, 5, 5, 6, 6, 6] + [6] * n_src_cols + [8, 7, 7, 6, 5, 30]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.freeze_panes = "A5"


def build_my_picks(wb):
    ws = wb.create_sheet("My Picks", 0)
    ws["A1"] = "Colin's 2026 Rookie Picks (linear draft, slot 11)"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws.merge_cells("A1:E1")
    picks = [
        (1, 11, 11, 11, "Own"),
        (2, 11, 23, 11, "Own"),
        (3, 9, 33, 9, "From CThompson1512 (Shadows of the Reaper)"),
        (3, 11, 35, 11, "Own"),
        (4, 11, 47, 11, "Own"),
    ]
    header_row(ws, 3, ["Round", "Pick", "Overall", "Slot", "Origin"])
    for i, row in enumerate(picks, 4):
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = NORMAL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center" if j != 5 else "left")
    for col, w in zip("ABCDE", [10, 10, 10, 10, 40]):
        ws.column_dimensions[col].width = w


def build_roster_tab(wb):
    """Tab showing your current Sleeper roster grouped by position with ages."""
    ws = wb.create_sheet("Roster")
    ws["A1"] = f"Current Roster — {LEAGUE_NAME}"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws.merge_cells("A1:E1")
    ws["A2"] = "Pulled from Sleeper API. Drives Roster Fit on the My Board tab."
    ws["A2"].font = NOTE
    ws.merge_cells("A2:E2")
    header_row(ws, 4, ["Pos", "Player", "NFL Team", "Age", "Yrs Exp"])
    # Hardcoded snapshot (refresh by re-running the bootstrap script)
    roster = [
        ("QB", "Lamar Jackson", "BAL", 29, 8),
        ("QB", "Brock Purdy", "SF", 26, 4),
        ("QB", "Tanner McKee", "PHI", 26, 3),
        ("QB", "Jaxson Dart", "NYG", 22, 1),
        ("RB", "Joe Mixon", "FA", 29, 9),
        ("RB", "Jonathan Taylor", "IND", 27, 6),
        ("RB", "Jerome Ford", "WAS", 26, 4),
        ("RB", "Trey Benson", "ARI", 23, 2),
        ("RB", "TreVeyon Henderson", "NE", 23, 1),
        ("RB", "Phil Mafah", "DAL", 23, 1),
        ("RB", "Kyle Monangai", "CHI", 23, 1),
        ("WR", "Michael Pittman", "PIT", 28, 6),
        ("WR", "DeVonta Smith", "PHI", 27, 5),
        ("WR", "Tank Dell", "HOU", 26, 3),
        ("WR", "George Pickens", "DAL", 25, 4),
        ("WR", "Ricky Pearsall", "SF", 25, 2),
        ("WR", "Drake London", "ATL", 24, 4),
        ("WR", "Jalen Coker", "CAR", 24, 2),
        ("WR", "Jack Bech", "LV", 23, 1),
        ("WR", "Troy Franklin", "DEN", 23, 2),
        ("WR", "Matthew Golden", "GB", 22, 1),
        ("TE", "Chig Okonkwo", "WAS", 26, 4),
        ("TE", "Kyle Pitts", "ATL", 25, 5),
        ("TE", "Tyler Warren", "IND", 23, 1),
        ("TE", "Harold Fannin", "CLE", 21, 1),
        ("LB", "Foyesade Oluokun", "JAX", 30, 8),
        ("LB", "Edgerrin Cooper", "GB", 24, 2),
        ("DB", "Brian Branch", "DET", 24, 3),
        ("DB", "Cooper DeJean", "PHI", 23, 2),
        ("K", "Ka'imi Fairbairn", "HOU", 32, 10),
        ("K", "Evan McPherson", "CIN", 26, 5),
    ]
    for i, (pos, name, team, age, yrs) in enumerate(roster, 5):
        for j, v in enumerate([pos, name, team, age, yrs], 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = NORMAL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center" if j != 2 else "left")
    # Needs summary
    base = 5 + len(roster) + 2
    ws.cell(row=base, column=1, value="Position Needs (drives Fit column)").font = SUBHEADER
    ws.cell(row=base, column=1).fill = SUBHEADER_FILL
    ws.merge_cells(start_row=base, start_column=1, end_row=base, end_column=5)
    needs_doc = [
        ("DL/EDGE", 1.0, "Zero DL on roster. Big hole."),
        ("WR", 0.7, "10 WRs but 5 are 25+. Need young top-end."),
        ("RB", 0.6, "Have Henderson; Mixon/JT aging."),
        ("LB", 0.6, "Cooper young; Oluokun aging."),
        ("TE", 0.5, "Warren+Fannin young. Marginal need."),
        ("CB/S", 0.3, "Branch + DeJean both young + good."),
        ("QB", 0.0, "4 deep. Skip."),
    ]
    header_row(ws, base + 1, ["Pos", "Need", "Reason", "", ""])
    for i, (pos, n, reason) in enumerate(needs_doc, base + 2):
        for j, v in enumerate([pos, n, reason, "", ""], 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = NORMAL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center" if j != 3 else "left")
    for col, w in zip("ABCDE", [8, 25, 12, 8, 50]):
        ws.column_dimensions[col].width = w


def build_sources_tab(wb, sources):
    ws = wb.create_sheet("Sources")
    ws["A1"] = "Source Registry & Ingest Workflow"
    ws["A1"].font = Font(name="Arial", bold=True, size=14)
    ws.merge_cells("A1:D1")
    header_row(ws, 3, ["Source key", "Label", "Weight", "Players loaded"])
    for i, (key, src) in enumerate(sources.items(), 4):
        row = [key, src["label"], src["weight"], len(src["ranks"])]
        for j, v in enumerate(row, 1):
            c = ws.cell(row=i, column=j, value=v)
            c.font = NORMAL
            c.border = BORDER
            c.alignment = Alignment(horizontal="center" if j != 2 else "left")
    base = 4 + len(sources) + 2
    instructions = [
        ("Workflow",
         "1. Drop a CSV in sources/ named e.g. dn.csv with columns: player,pos,team,source_rank\n"
         "2. Add an entry in sources/sources.json with key, label, csv, weight.\n"
         "3. Run: python build_xlsx_v2.py"),
        ("Live Sleeper ADP",
         "Once your Phoenix League draft starts (~May 4), run sleeper_scraper.py with "
         "your league's draft_id (1335118685605498880) AND a few mock IDs you've joined. "
         "Output goes to sources/sleeper_adp.csv. Already wired in sources.json."),
        ("FantasyCalc refresh",
         "python fc_fetcher.py — hits api.fantasycalc.com (no auth). "
         "Filter is by player.maybeYoe == 0 (rookie). "
         "Note: FC has no public TEP toggle — values are non-TEP."),
        ("Roster Fit",
         "Driven by ROSTER_NEED in build_xlsx_v2.py. Edit that dict to retune."
         " Adjusts up for young (<=21), down for old (>=24). DL/EDGE = priority "
         "since Phoenix League roster has zero defensive linemen."),
        ("Capital Score",
         "Lower = better. = (NFL pick * round multiplier) + position bump. "
         "Bumps: QB=-18, TE=-10, RB=-6, WR=0, DL/EDGE=-12, LB=-8, DB/S/CB=-6."),
    ]
    for i, (k, v) in enumerate(instructions):
        row = base + i
        c1 = ws.cell(row=row, column=1, value=k)
        c1.font = SUBHEADER
        c1.fill = SUBHEADER_FILL
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c2 = ws.cell(row=row, column=3, value=v)
        c2.font = NORMAL
        c2.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
        ws.row_dimensions[row].height = max(40, 14 * (v.count("\n") + 2))
    for col, w in zip("ABCDEFGHIJ", [16, 30, 9, 12, 12, 12, 12, 12, 12, 12]):
        ws.column_dimensions[col].width = w


# ============= DOCS JSON EXPORT =============

def write_docs_json(ranked, sources):
    """Emit docs/data.json — the rankings the static webpage consumes."""
    src_keys = list(sources.keys())
    src_meta = [
        {"key": k, "label": sources[k]["label"], "weight": sources[k]["weight"],
         "n_players": len(sources[k]["ranks"])}
        for k in src_keys
    ]
    players_out = []
    for p in ranked:
        ranks_present = list(p["source_ranks"].values())
        spread = max(ranks_present) - min(ranks_present) if len(ranks_present) >= 2 else None
        ktc_rk = p["source_ranks"].get("ktc")
        notes = []
        if p["pos"] == "TE":
            notes.append("TEP")
        if p["pos"] == "QB":
            notes.append("SF")
        if p["pos"] in ("DL", "EDGE", "LB", "DB", "S", "CB"):
            notes.append("IDP")
        if p.get("rd") == 1:
            notes.append("R1 cap")
        if p.get("rd") and p["rd"] >= 6:
            notes.append("late dart")
        if not p["source_ranks"]:
            notes.append("no source")
        if ktc_rk is not None and abs(ktc_rk - p["cap_rk"]) >= 8:
            notes.append("KTC loves" if ktc_rk < p["cap_rk"] else "KTC fades")
        if spread is not None and spread >= 15:
            notes.append(f"high spread ({spread})")
        fit_mark, fit_label = roster_fit(p)
        slot = PICK_AT_RANK.get(p["my_rk"], "")

        players_out.append({
            "my_rk": p["my_rk"],
            "tier": p["tier"],
            "slot": slot,
            "name": p["name"],
            "pos": p["pos"],
            "team": p["team"] or "FA",
            "age": p.get("age"),
            "rd": p.get("rd"),
            "pk": p.get("pk"),
            "college": p.get("college") or "",
            "sleeper_id": p.get("sleeper_id"),
            "source_ranks": p["source_ranks"],
            "ktc_value": p["source_values"].get("ktc"),
            "fc_value": p["source_values"].get("fc"),
            "cap_rk": p["cap_rk"],
            "spread": spread,
            "n_sources": p.get("n_sources", 0),
            "fit_mark": fit_mark,
            "fit_label": fit_label,
            "final": round(p["final"], 2),
            "notes": notes,
        })

    out = {
        "league": {
            "name": LEAGUE_NAME,
            "format": LEAGUE_FORMAT,
            "league_id": "1335118685588701184",
            "draft_id": "1335118685605498880",
        },
        "owner": {
            "username": "colindiggs",
            "user_id": "1250202714625822720",
            "roster_id": 3,
            "picks": [
                {"round": 1, "pick": 11, "overall": 11, "origin": "Own"},
                {"round": 2, "pick": 11, "overall": 23, "origin": "Own"},
                {"round": 3, "pick": 9,  "overall": 33, "origin": "From CThompson1512 (Shadows of the Reaper)"},
                {"round": 3, "pick": 11, "overall": 35, "origin": "Own"},
                {"round": 4, "pick": 11, "overall": 47, "origin": "Own"},
            ],
        },
        "weights": {k: sources[k]["weight"] for k in src_keys} | {"cap": CAP_WEIGHT},
        "roster_need": ROSTER_NEED,
        "sources": src_meta,
        "players": players_out,
    }

    os.makedirs(DOCS_DIR, exist_ok=True)
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"Saved {OUT_JSON} ({len(players_out)} players)")


# ============= MAIN =============

def main():
    print("Loading Sleeper player index...")
    sleeper_idx = load_sleeper_idx()
    print(f"  -> {len(sleeper_idx)} players cached")

    print("\nLoading sources...")
    sources = load_sources()
    if not sources:
        print("No sources loaded.")
        return

    print(f"\nBuilding player list...")
    players = build_players(sources, sleeper_idx)
    print(f"  -> {len(players)} unique players")
    print("Computing weighted blend...")
    ranked = compute_blend(players, sources)

    print(f"\nTop 25 by My Rk ({len(sources)} sources):")
    src_keys = list(sources.keys())
    print(f"  {'Rk':>3}. {'Player':24s} {'Pos':3s} {'Team':4s} {'Age':3s} | "
          + " ".join(f"{k[:4]:>4}" for k in src_keys) + " | cap")
    for p in ranked[:25]:
        cells = []
        for k in src_keys:
            r = p["source_ranks"].get(k)
            cells.append(f"{r:>4}" if r is not None else "   -")
        age = p.get("age") or "-"
        print(f"  {p['my_rk']:>3}. {p['name']:24s} {p['pos']:3s} {p['team']:4s} {str(age):>3} | "
              f"{' '.join(cells)} | {p['cap_rk']:>3}")

    wb = Workbook()
    wb.remove(wb.active)
    build_my_picks(wb)
    build_my_board(wb, ranked, sources)
    build_roster_tab(wb)
    build_sources_tab(wb, sources)
    wb.save(OUT_XLSX)
    print(f"\nSaved {OUT_XLSX}")

    write_docs_json(ranked, sources)


if __name__ == "__main__":
    main()
